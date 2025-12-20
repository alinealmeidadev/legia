"""
LEGIA PLATFORM - Serviço de Importação Excel
Processa importação em lote de clientes via Excel
"""
import pandas as pd
from typing import List, Dict, Any
from io import BytesIO
import openpyxl
from openpyxl import Workbook


class ExcelService:
    """Serviço para importação de clientes via Excel"""

    # Colunas esperadas no Excel
    COLUNAS_TEMPLATE = [
        "tipo",  # pf ou pj
        "nome",
        "documento",  # CPF ou CNPJ
        "email",
        "telefone",
        "celular",
        "razao_social",  # Apenas PJ
        "nome_fantasia",  # Apenas PJ
        "cnae_principal",  # Apenas PJ
        "inscricao_estadual",  # Apenas PJ
        "inscricao_municipal",  # Apenas PJ
        "rua",
        "numero",
        "complemento",
        "bairro",
        "cidade",
        "estado",
        "cep"
    ]

    @staticmethod
    def gerar_template_excel() -> bytes:
        """
        Gera arquivo Excel template para importação

        Returns:
            Bytes do arquivo Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Cabeçalhos
        for i, coluna in enumerate(ExcelService.COLUNAS_TEMPLATE, start=1):
            ws.cell(row=1, column=i, value=coluna)

        # Adicionar linha de exemplo
        exemplo = [
            "pj",
            "Empresa Exemplo Ltda",
            "12.345.678/0001-90",
            "contato@exemplo.com",
            "(11) 1234-5678",
            "(11) 91234-5678",
            "Empresa Exemplo Ltda",
            "Exemplo",
            "6201-5/00",
            "123456789",
            "987654321",
            "Rua das Flores",
            "123",
            "Sala 45",
            "Centro",
            "São Paulo",
            "SP",
            "01000-000"
        ]

        for i, valor in enumerate(exemplo, start=1):
            ws.cell(row=2, column=i, value=valor)

        # Salvar em bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    @staticmethod
    async def processar_excel(
        arquivo_bytes: bytes,
        schema_name: str
    ) -> Dict[str, Any]:
        """
        Processa arquivo Excel e valida dados

        Args:
            arquivo_bytes: Bytes do arquivo Excel
            schema_name: Schema do tenant

        Returns:
            Dict com resultado do processamento
        """
        try:
            # Ler Excel
            buffer = BytesIO(arquivo_bytes)
            df = pd.read_excel(buffer)

            # Verificar colunas obrigatórias
            colunas_obrigatorias = ["tipo", "nome", "documento"]
            colunas_faltando = [col for col in colunas_obrigatorias if col not in df.columns]

            if colunas_faltando:
                return {
                    "success": False,
                    "error": f"Colunas obrigatórias faltando: {', '.join(colunas_faltando)}"
                }

            # Processar linhas
            clientes_validos = []
            erros = []

            for index, row in df.iterrows():
                linha_numero = index + 2  # +2 porque Excel começa em 1 e temos header

                try:
                    # Validar tipo
                    tipo = str(row.get('tipo', '')).strip().lower()
                    if tipo not in ['pf', 'pj']:
                        erros.append(f"Linha {linha_numero}: Tipo inválido (deve ser 'pf' ou 'pj')")
                        continue

                    # Validar campos obrigatórios
                    nome = str(row.get('nome', '')).strip()
                    documento = str(row.get('documento', '')).strip()

                    if not nome or nome == 'nan':
                        erros.append(f"Linha {linha_numero}: Nome é obrigatório")
                        continue

                    if not documento or documento == 'nan':
                        erros.append(f"Linha {linha_numero}: Documento é obrigatório")
                        continue

                    # Montar cliente
                    cliente = {
                        "type": tipo,
                        "name": nome,
                        "document": documento,
                        "email": str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
                        "phone": str(row.get('telefone', '')).strip() if pd.notna(row.get('telefone')) else None,
                        "mobile": str(row.get('celular', '')).strip() if pd.notna(row.get('celular')) else None,
                        "address_street": str(row.get('rua', '')).strip() if pd.notna(row.get('rua')) else None,
                        "address_number": str(row.get('numero', '')).strip() if pd.notna(row.get('numero')) else None,
                        "address_complement": str(row.get('complemento', '')).strip() if pd.notna(row.get('complemento')) else None,
                        "address_neighborhood": str(row.get('bairro', '')).strip() if pd.notna(row.get('bairro')) else None,
                        "address_city": str(row.get('cidade', '')).strip() if pd.notna(row.get('cidade')) else None,
                        "address_state": str(row.get('estado', '')).strip() if pd.notna(row.get('estado')) else None,
                        "address_zipcode": str(row.get('cep', '')).strip() if pd.notna(row.get('cep')) else None,
                    }

                    # Dados PJ
                    if tipo == 'pj':
                        cliente.update({
                            "company_name": str(row.get('razao_social', '')).strip() if pd.notna(row.get('razao_social')) else None,
                            "trade_name": str(row.get('nome_fantasia', '')).strip() if pd.notna(row.get('nome_fantasia')) else None,
                            "cnae_primary": str(row.get('cnae_principal', '')).strip() if pd.notna(row.get('cnae_principal')) else None,
                            "state_registration": str(row.get('inscricao_estadual', '')).strip() if pd.notna(row.get('inscricao_estadual')) else None,
                            "municipal_registration": str(row.get('inscricao_municipal', '')).strip() if pd.notna(row.get('inscricao_municipal')) else None,
                        })

                    clientes_validos.append({
                        "linha": linha_numero,
                        "dados": cliente
                    })

                except Exception as e:
                    erros.append(f"Linha {linha_numero}: Erro ao processar - {str(e)}")

            return {
                "success": True,
                "total_linhas": len(df),
                "clientes_validos": clientes_validos,
                "total_validos": len(clientes_validos),
                "erros": erros,
                "total_erros": len(erros)
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"Erro ao processar Excel: {str(e)}"
            }
